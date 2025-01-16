
###############################################
## COPIED FROM IIRO 'sp-talytics-task1' REPO
###############################################

from openpyxl import load_workbook

#This is the database schema that showcases the structure of the database
'''{
    "user_id": "string",  // Unique identifier for the user
    "speed_data": [
        {
            "speedzone": "string",      // e.g., "0-3.4"
            "total_time": "string",    // e.g., "00:33:51"
            "time_percentage": "float", // e.g., 62.76
            "total_distance": "float",  // e.g., 620.05
            "distance_percentage": "float" // e.g., 34.34
        },
        ... Here will be another similar block with a different speedzone e.g "3.5-9.9"
    ]
}'''
def get_user_data(full_path):
    #This is the sample user data so that the true data can be appended to it.
    user_data = []
    #Next we will pull the data from the excel file and format it to the schema.

    #Define data_file as the path to file !!This must be made generic so that it can handle multiple data files!!
    data_file = full_path 

    #Load entire workbook aka all work sheets in the excel file
    workbook = load_workbook(data_file)
    #this variable is to control the iteration of the worksheets
    worksheet_count = 0
    #this BIG for-loop loops through all the worksheets e.g Visa pelaaja, Veeti pelaaja
    for sheet in workbook.worksheets:
    #Select active sheet e.g the first sheet, this one has Visa pelaaja 
        active_sheet = sheet
        #create the first instance of the user data dictionary and give the user an id.
        user_data.append({
            'user_id' : sheet.title,
            'speed_data' : [],
            'acceleration': 0,
            'deceleration': 0
        })
        #col_names contains the titles of the needed data, this allows the iteration of those columns
        col_names = {}
        current = 0

        #Iterate through the whole sheet to create a dictionary with column names and their column number.
        #E.g speedzone is column 8, this will be used later to get all relevant data.
        for COL in active_sheet.iter_cols(1, active_sheet.max_column):
            #print(COL[0].value)
            if(COL[0].value == None and current > 5):
                col_names["None2"] = current
            elif(COL[0].value == "Percentage (%)" and current > 10):
                col_names["Percentage2"] = current
            else:
                col_names[COL[0].value] = current
            current += 1
        #print(col_names)

        #Now iterate through the table rows to find all rows that contain data from the chosen column.
        for row_cells in active_sheet.iter_rows(min_row=2, max_row=5):
            if(row_cells[col_names['speedzone']].value != None and row_cells[col_names['speedzone']].value != 'speedzone'):
                # Create a new dictionary for each row with data, we can use this dictionary now to add the rest of the values.
                speed_data_entry = {
                    "speedzone": row_cells[col_names['speedzone']].value,
                    "total_time": "",
                    "time_percentage": 0.0,
                    "total_distance": 0.0,
                    "distance_percentage": 0.0,
                    "acceleration": 0,
                    "decelaration": 0
                }
                # Append the dictionary to the list
                user_data[worksheet_count]["speed_data"].append(speed_data_entry)

        #Now just append the rest of the values to the dictionary.

        #BUNCH OF USELESS REPETITION, but works.
        #total_time
        iteration = 0
        for row_cells in active_sheet.iter_rows(min_row=2, max_row=5):
            if(row_cells[col_names['Total Time']].value != None):
                user_data[worksheet_count]["speed_data"][iteration]["total_time"] = row_cells[col_names['Total Time']].value
                iteration += 1
        #time_percentage
        iteration = 0
        for row_cells in active_sheet.iter_rows(min_row=2, max_row=5):
            if(row_cells[col_names['Percentage (%)']].value != None):
                user_data[worksheet_count]["speed_data"][iteration]["time_percentage"] = float(row_cells[col_names['Percentage (%)']].value)
                iteration += 1
        #total_distance
        iteration = 0
        for row_cells in active_sheet.iter_rows(min_row=2, max_row=5):
            if(row_cells[col_names['Total Distance (m)']].value != None):
                user_data[worksheet_count]["speed_data"][iteration]["total_distance"] = float(row_cells[col_names['Total Distance (m)']].value)
                iteration += 1
        #distance_percentage
        iteration = 0
        for row_cells in active_sheet.iter_rows(min_row=2, max_row=5):
            if(row_cells[col_names['Percentage2']].value != None):
                user_data[worksheet_count]["speed_data"][iteration]["distance_percentage"] = float(row_cells[col_names['Percentage2']].value)
                iteration += 1
        #acceleration
        iteration = 0
        for row_cells in active_sheet.iter_rows(min_row=12, max_row=13):
            if(row_cells[col_names['Total Time']].value != None):
                user_data[worksheet_count]["acceleration"] = int(row_cells[col_names['Total Time']].value)
                iteration += 1
        #deceleration
        iteration = 0
        for row_cells in active_sheet.iter_rows(min_row=12, max_row=13):
            if(row_cells[col_names['Percentage (%)']].value != None):
                user_data[worksheet_count]['deceleration'] = row_cells[col_names['Percentage (%)']].value
                iteration += 1

        worksheet_count += 1
    return(user_data)